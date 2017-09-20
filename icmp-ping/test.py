import ping
import openpyxl as pyxl
import pandas as pd
from openpyxl.utils.dataframe import dataframe_to_rows

# ping.verbose_ping(hostname="www.ruc.edu.cn", packet_size=32)
# ping.verbose_ping(hostname="www.zuel.edu.cn", packet_size=32)
# ping.verbose_ping(hostname="www.cupes.edu.cn", packet_size=32)

inputsheet = pyxl.load_workbook("web_list.xlsx")
inputsheet.template = True
worksheet = inputsheet.active


webframe = pd.DataFrame(worksheet.values)

count=0
for row in dataframe_to_rows(webframe, index=True, header=False):
    count+=1
    if worksheet.cell(row=count,column=3).value:
        print(row)
        school=worksheet.cell(row=count,column=2).value
        url=worksheet.cell(row=count,column=3).value
        packet_loss,minimum,average,maximum=ping.verbose_ping(hostname=url)
        # print(str(school)+"   "+str(url)+"   "+"round-trip (ms)   min/avg/max = %d/%d/%d   packet loss=%0.1f " % (
        #         minimum, average, maximum,packet_loss
        #     ))
        worksheet.cell(row=count,column=4).value=minimum
        worksheet.cell(row=count,column=5).value=average
        worksheet.cell(row=count,column=6).value=maximum
        worksheet.cell(row=count,column=7).value=packet_loss

inputsheet.save('template.xltx')