from __future__ import print_function, unicode_literals
import sys
import openpyxl as pyxl
import pandas as pd
from openpyxl.utils.dataframe import dataframe_to_rows
import jieba
jieba.load_userdict("userdict.txt")
import jieba.posseg as pseg

inputsheet = pyxl.load_workbook('excel\\target.xlsx')
inputsheet.template = True
worksheet = inputsheet.active

targetexcel = pyxl.Workbook()
targetsheet = targetexcel.active

# for j in range(1, 8):
#     data = worksheet[.cell(row=1, column=j)].value
#     print(data)

# def debit(i, columnindex):
#     worksheet.cell(row=i, column=columnindex).value = '借:' + \
#         worksheet.cell(row=i, column=columnindex).value
#     return True

# def credit(i, columnindex):
#     worksheet.cell(row=i, column=columnindex).value = '贷:' + \
#         worksheet.cell(row=i, column=columnindex).value
#     return True

# for i in range(2, 199):
#     if not i % 2:
#         # i=2
#         debit(i, 2)
#         credit(i + 1, 2)
#         worksheet.cell(row=i, column=5).value = worksheet.cell(
#             row=i + 1, column=2).value
#         worksheet.cell(row=i, column=7).value = worksheet.cell(
#             row=i + 1, column=7).value

# inputsheet.save('document_template.xltx')

df = pd.DataFrame(worksheet.values)

for r in dataframe_to_rows(df, index=True, header=True):
    if type(r[2]) == str:
        entry = '<BOS> ' + (' '.join(jieba.cut(r[2]))) + ' <EOS>'
        # debit_project = (' '.join(jieba.cut(r[3])))
        # credit_project = (' '.join(jieba.cut(r[6])))
        cop = r[3] + str(r[5]) + r[6] + str(r[8])
        answer = '<BOS> ' + (' '.join(jieba.cut(cop))) + ' <EOS>'
        row = [r[1], entry, answer]
        # r[2] = entry
        # r[3] = debit_project
        # r[6] = credit_project
        print(row)
        targetsheet.append(row)


targetexcel.save('entry-answer.xlsx')
