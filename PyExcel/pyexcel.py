import openpyxl as pyxl
import pandas as pd
from openpyxl.utils.dataframe import dataframe_to_rows

inputsheet = pyxl.load_workbook('template.xlsx')
inputsheet.template = True
worksheet = inputsheet.active

targetexcel = pyxl.Workbook()
targetsheet = targetexcel.active

# for j in range(1, 8):
#     data = worksheet[.cell(row=1, column=j)].value
#     print(data)

# def debit(i, columnindex):
#     worksheet.cell(row=i, column=columnindex).value = '借:' + \
#         worksheet.cell(row=i, column=columnindex).value
#     return True

# def credit(i, columnindex):
#     worksheet.cell(row=i, column=columnindex).value = '贷:' + \
#         worksheet.cell(row=i, column=columnindex).value
#     return True

# for i in range(2, 199):
#     if not i % 2:
#         # i=2
#         debit(i, 2)
#         credit(i + 1, 2)
#         worksheet.cell(row=i, column=5).value = worksheet.cell(
#             row=i + 1, column=2).value
#         worksheet.cell(row=i, column=7).value = worksheet.cell(
#             row=i + 1, column=7).value

# inputsheet.save('document_template.xltx')

df = pd.DataFrame(worksheet.values)

for r in dataframe_to_rows(df, index=True, header=True):
    if r[1]:
        targetsheet.append(r)
        print(r)


targetexcel.save('target.xlsx')
